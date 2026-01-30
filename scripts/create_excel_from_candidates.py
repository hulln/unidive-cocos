#!/usr/bin/env python3
"""
Convert backchannel_candidates.csv to a formatted Excel file with color coding.
"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def main():
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    
    csv_path = project_root / "output" / "sst" / "backchannel_candidates.csv"
    xlsx_path = project_root / "output" / "sst" / "backchannel_candidates.xlsx"
    
    # Read CSV
    with csv_path.open("r", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Backchannel Candidates"
    
    # Write header with bold font
    header = rows[0]
    for col_idx, value in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.font = Font(bold=True)
    
    # Color fills for flags
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # A_looks_like_backchannel
    red_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")     # B_has_content
    orange_fill = PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")  # B_is_question
    pink_fill = PatternFill(start_color="FFE6F0", end_color="FFE6F0", fill_type="solid")    # B_after_question
    
    # Confidence colors
    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")    # GREEN
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # YELLOW
    low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # LIGHT RED
    
    # Find column indices
    confidence_col = header.index("confidence") + 1
    a_looks_col = header.index("A_looks_like_backchannel") + 1
    b_has_content_col = header.index("B_has_content") + 1
    b_is_question_col = header.index("B_is_question") + 1
    b_after_question_col = header.index("B_after_question") + 1
    
    # Write data rows with conditional formatting
    for row_idx, row_data in enumerate(rows[1:], 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Color code confidence column
            if col_idx == confidence_col:
                if value == "HIGH":
                    cell.fill = high_fill
                elif value == "MEDIUM":
                    cell.fill = medium_fill
                elif value == "LOW":
                    cell.fill = low_fill
            
            # Color code flag columns
            if col_idx == a_looks_col and value == "1":
                cell.fill = yellow_fill
            elif col_idx == b_has_content_col and value == "1":
                cell.fill = red_fill
            elif col_idx == b_is_question_col and value == "1":
                cell.fill = orange_fill
            elif col_idx == b_after_question_col and value == "1":
                cell.fill = pink_fill
    
    # Auto-adjust column widths
    for col_idx in range(1, len(header) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws[col_letter]:
            try:
                cell_length = len(str(row.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 80)  # cap at 80
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    wb.save(xlsx_path)
    print(f"Created Excel file: {xlsx_path}")
    print(f"  - HIGH confidence (green): best candidates")
    print(f"  - MEDIUM confidence (yellow): good candidates")  
    print(f"  - LOW confidence (light red): review carefully")
    print(f"  - Yellow flag: A looks like backchannel (weird context)")
    print(f"  - Red flag: B has content words (VERB/NOUN/ADJ)")
    print(f"  - Orange flag: B is multi-token question")
    print(f"  - Pink flag: B comes after A's question (could be answer)")

if __name__ == "__main__":
    main()
